"""
Excel 配置文件导入导出功能
支持设备配置的导入和导出
"""
from typing import List, Dict, Any
import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .device_manager import DeviceConfig, ConnectionType, OperationResult


class ExcelManager:
    """Excel 配置管理器"""
    
    # Excel 列定义
    COLUMNS = [
        '设备ID', '设备名称', '连接类型', '从站地址',
        '串口端口', '波特率', '数据位', '校验位', '停止位',
        'IP地址', 'TCP端口', '超时时间(秒)'
    ]
    
    def __init__(self):
        """初始化 Excel 管理器"""
        pass
    
    def export_devices(self, devices: List[DeviceConfig], file_path: str) -> OperationResult:
        """
        导出设备配置到 Excel 文件
        
        Args:
            devices: 设备配置列表
            file_path: Excel 文件路径
            
        Returns:
            操作结果
        """
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "设备配置"
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 写入表头
            for col_idx, col_name in enumerate(self.COLUMNS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 写入数据
            for row_idx, device in enumerate(devices, start=2):
                ws.cell(row=row_idx, column=1, value=device.device_id)
                ws.cell(row=row_idx, column=2, value=device.name)
                ws.cell(row=row_idx, column=3, value=device.connection_type.value)
                ws.cell(row=row_idx, column=4, value=device.slave_id)
                
                # RTU 配置
                ws.cell(row=row_idx, column=5, value=device.port or "")
                ws.cell(row=row_idx, column=6, value=device.baudrate)
                ws.cell(row=row_idx, column=7, value=device.bytesize)
                ws.cell(row=row_idx, column=8, value=device.parity)
                ws.cell(row=row_idx, column=9, value=device.stopbits)
                
                # TCP 配置
                ws.cell(row=row_idx, column=10, value=device.host or "")
                ws.cell(row=row_idx, column=11, value=device.tcp_port)
                
                # 通用配置
                ws.cell(row=row_idx, column=12, value=device.timeout)
            
            # 自动调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # 保存文件
            wb.save(file_path)
            return OperationResult(True, data=f"成功导出 {len(devices)} 个设备配置")
            
        except Exception as e:
            return OperationResult(False, error=f"导出失败: {str(e)}")
    
    def import_devices(self, file_path: str) -> OperationResult:
        """
        从 Excel 文件导入设备配置
        
        Args:
            file_path: Excel 文件路径
            
        Returns:
            操作结果，成功时 data 包含设备配置列表
        """
        try:
            # 检查文件是否存在
            if not Path(file_path).exists():
                return OperationResult(False, error="文件不存在")
            
            # 读取 Excel 文件
            df = pd.read_excel(file_path, sheet_name="设备配置")
            
            # 验证列名
            required_columns = self.COLUMNS
            missing_columns = set(required_columns) - set(df.columns)
            if missing_columns:
                return OperationResult(
                    False, 
                    error=f"Excel 文件缺少必需的列: {', '.join(missing_columns)}"
                )
            
            devices = []
            errors = []
            
            # 解析每一行
            for idx, row in df.iterrows():
                try:
                    # 解析连接类型
                    conn_type_str = str(row['连接类型']).strip().upper()
                    if conn_type_str == "RTU":
                        connection_type = ConnectionType.RTU
                    elif conn_type_str == "TCP":
                        connection_type = ConnectionType.TCP
                    else:
                        errors.append(f"第 {idx+2} 行: 无效的连接类型 '{conn_type_str}'")
                        continue
                    
                    # 创建设备配置
                    config = DeviceConfig(
                        device_id=str(row['设备ID']).strip(),
                        name=str(row['设备名称']).strip(),
                        connection_type=connection_type,
                        slave_id=int(row['从站地址']),
                        timeout=float(row['超时时间(秒)']) if pd.notna(row['超时时间(秒)']) else 3.0
                    )
                    
                    # RTU 配置
                    if connection_type == ConnectionType.RTU:
                        config.port = str(row['串口端口']).strip() if pd.notna(row['串口端口']) else None
                        config.baudrate = int(row['波特率']) if pd.notna(row['波特率']) else 9600
                        config.bytesize = int(row['数据位']) if pd.notna(row['数据位']) else 8
                        config.parity = str(row['校验位']).strip() if pd.notna(row['校验位']) else 'N'
                        config.stopbits = int(row['停止位']) if pd.notna(row['停止位']) else 1
                        
                        if not config.port:
                            errors.append(f"第 {idx+2} 行: RTU 设备缺少串口端口")
                            continue
                    
                    # TCP 配置
                    elif connection_type == ConnectionType.TCP:
                        config.host = str(row['IP地址']).strip() if pd.notna(row['IP地址']) else None
                        config.tcp_port = int(row['TCP端口']) if pd.notna(row['TCP端口']) else 502
                        
                        if not config.host:
                            errors.append(f"第 {idx+2} 行: TCP 设备缺少IP地址")
                            continue
                    
                    devices.append(config)
                    
                except Exception as e:
                    errors.append(f"第 {idx+2} 行解析错误: {str(e)}")
            
            if errors:
                error_msg = "\n".join(errors)
                if not devices:
                    return OperationResult(False, error=f"导入失败:\n{error_msg}")
                else:
                    return OperationResult(
                        True, 
                        data=devices,
                        error=f"部分行导入失败:\n{error_msg}"
                    )
            
            return OperationResult(True, data=devices)
            
        except Exception as e:
            return OperationResult(False, error=f"导入失败: {str(e)}")
    
    def create_template(self, file_path: str) -> OperationResult:
        """
        创建配置模板文件
        
        Args:
            file_path: 模板文件保存路径
            
        Returns:
            操作结果
        """
        try:
            # 创建示例设备配置
            sample_devices = [
                DeviceConfig(
                    device_id="RTU_001",
                    name="RTU测试设备",
                    connection_type=ConnectionType.RTU,
                    slave_id=1,
                    port="COM3",
                    baudrate=9600,
                    bytesize=8,
                    parity='N',
                    stopbits=1,
                    timeout=3.0
                ),
                DeviceConfig(
                    device_id="TCP_001",
                    name="TCP测试设备",
                    connection_type=ConnectionType.TCP,
                    slave_id=1,
                    host="192.168.1.100",
                    tcp_port=502,
                    timeout=3.0
                )
            ]
            
            return self.export_devices(sample_devices, file_path)
            
        except Exception as e:
            return OperationResult(False, error=f"创建模板失败: {str(e)}")
    
    def export_operation_log(self, log_data: List[Dict[str, Any]], 
                            file_path: str) -> OperationResult:
        """
        导出操作日志到 Excel
        
        Args:
            log_data: 日志数据列表
            file_path: 保存路径
            
        Returns:
            操作结果
        """
        try:
            df = pd.DataFrame(log_data)
            
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='操作日志', index=False)
                
                # 获取工作表
                workbook = writer.book
                worksheet = writer.sheets['操作日志']
                
                # 设置表头样式
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # 自动调整列宽
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            return OperationResult(True, data=f"成功导出 {len(log_data)} 条日志")
            
        except Exception as e:
            return OperationResult(False, error=f"导出日志失败: {str(e)}")
