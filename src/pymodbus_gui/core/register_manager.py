"""
寄存器点表管理器
负责管理 Modbus Slave 的寄存器点表配置
"""
from typing import List, Dict, Optional
import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .slave_server import RegisterPoint, OperationResult


class RegisterManager:
    """寄存器点表管理器"""
    
    # Excel 列定义
    COLUMNS = [
        '地址', '点位名称', '寄存器类型', '初始值', '描述',
        '单位', '最小值', '最大值', '只读'
    ]
    
    # 寄存器类型映射
    REGISTER_TYPES = {
        '线圈': 'coil',
        '离散输入': 'discrete_input',
        '保持寄存器': 'holding_register',
        '输入寄存器': 'input_register',
        'coil': 'coil',
        'discrete_input': 'discrete_input',
        'holding_register': 'holding_register',
        'input_register': 'input_register'
    }
    
    REGISTER_TYPES_CN = {
        'coil': '线圈',
        'discrete_input': '离散输入',
        'holding_register': '保持寄存器',
        'input_register': '输入寄存器'
    }
    
    def __init__(self):
        """初始化寄存器管理器"""
        pass
    
    def export_register_points(self, points: List[RegisterPoint], file_path: str) -> OperationResult:
        """
        导出寄存器点表到 Excel
        
        Args:
            points: 寄存器点位列表
            file_path: Excel 文件路径
            
        Returns:
            操作结果
        """
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "寄存器点表"
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 写入表头
            for col_idx, col_name in enumerate(self.COLUMNS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 写入数据
            for row_idx, point in enumerate(points, start=2):
                ws.cell(row=row_idx, column=1, value=point.address)
                ws.cell(row=row_idx, column=2, value=point.name)
                ws.cell(row=row_idx, column=3, value=self.REGISTER_TYPES_CN.get(point.register_type, point.register_type))
                ws.cell(row=row_idx, column=4, value=point.value)
                ws.cell(row=row_idx, column=5, value=point.description)
                ws.cell(row=row_idx, column=6, value=point.unit)
                ws.cell(row=row_idx, column=7, value=point.min_value if point.min_value is not None else "")
                ws.cell(row=row_idx, column=8, value=point.max_value if point.max_value is not None else "")
                ws.cell(row=row_idx, column=9, value="是" if point.read_only else "否")
            
            # 自动调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # 保存文件
            wb.save(file_path)
            return OperationResult(True, data=f"成功导出 {len(points)} 个寄存器点位")
            
        except Exception as e:
            return OperationResult(False, error=f"导出失败: {str(e)}")
    
    def import_register_points(self, file_path: str) -> OperationResult:
        """
        从 Excel 导入寄存器点表
        
        Args:
            file_path: Excel 文件路径
            
        Returns:
            操作结果，成功时 data 包含寄存器点位列表
        """
        try:
            # 检查文件是否存在
            if not Path(file_path).exists():
                return OperationResult(False, error="文件不存在")
            
            # 读取 Excel 文件
            df = pd.read_excel(file_path, sheet_name="寄存器点表")
            
            # 验证列名
            required_columns = self.COLUMNS
            missing_columns = set(required_columns) - set(df.columns)
            if missing_columns:
                return OperationResult(
                    False, 
                    error=f"Excel 文件缺少必需的列: {', '.join(missing_columns)}"
                )
            
            points = []
            errors = []
            
            # 解析每一行
            for idx, row in df.iterrows():
                try:
                    # 解析寄存器类型
                    reg_type_str = str(row['寄存器类型']).strip()
                    register_type = self.REGISTER_TYPES.get(reg_type_str)
                    
                    if not register_type:
                        errors.append(f"第 {idx+2} 行: 无效的寄存器类型 '{reg_type_str}'")
                        continue
                    
                    # 解析值
                    value = row['初始值']
                    if pd.isna(value):
                        value = 0
                    elif register_type in ['coil', 'discrete_input']:
                        # 布尔类型处理
                        if isinstance(value, str):
                            value = 1 if value.lower() in ['true', '是', '1', 'on'] else 0
                        else:
                            value = int(value)
                    else:
                        value = int(value)
                    
                    # 解析只读
                    read_only = False
                    if pd.notna(row['只读']):
                        read_only_str = str(row['只读']).strip().lower()
                        read_only = read_only_str in ['true', '是', '1', 'yes']
                    
                    # 创建点位配置
                    point = RegisterPoint(
                        address=int(row['地址']),
                        name=str(row['点位名称']).strip(),
                        register_type=register_type,
                        value=value,
                        description=str(row['描述']).strip() if pd.notna(row['描述']) else "",
                        unit=str(row['单位']).strip() if pd.notna(row['单位']) else "",
                        min_value=float(row['最小值']) if pd.notna(row['最小值']) else None,
                        max_value=float(row['最大值']) if pd.notna(row['最大值']) else None,
                        read_only=read_only
                    )
                    
                    points.append(point)
                    
                except Exception as e:
                    errors.append(f"第 {idx+2} 行解析错误: {str(e)}")
            
            if errors:
                error_msg = "\n".join(errors)
                if not points:
                    return OperationResult(False, error=f"导入失败:\n{error_msg}")
                else:
                    return OperationResult(
                        True, 
                        data=points,
                        error=f"部分行导入失败:\n{error_msg}"
                    )
            
            return OperationResult(True, data=points)
            
        except Exception as e:
            return OperationResult(False, error=f"导入失败: {str(e)}")
    
    def create_template(self, file_path: str) -> OperationResult:
        """
        创建寄存器点表模板
        
        Args:
            file_path: 模板文件保存路径
            
        Returns:
            操作结果
        """
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "寄存器点表"
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 写入表头
            for col_idx, col_name in enumerate(self.COLUMNS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 写入示例数据
            examples = [
                {
                    '地址': 0,
                    '点位名称': '运行状态',
                    '寄存器类型': '线圈',
                    '初始值': 0,
                    '描述': '设备运行状态',
                    '单位': '',
                    '最小值': '',
                    '最大值': '',
                    '只读': '否'
                },
                {
                    '地址': 1,
                    '点位名称': '故障报警',
                    '寄存器类型': '离散输入',
                    '初始值': 0,
                    '描述': '设备故障报警',
                    '单位': '',
                    '最小值': '',
                    '最大值': '',
                    '只读': '是'
                },
                {
                    '地址': 0,
                    '点位名称': '温度设定',
                    '寄存器类型': '保持寄存器',
                    '初始值': 250,
                    '描述': '温度设定值',
                    '单位': '0.1°C',
                    '最小值': 0,
                    '最大值': 1000,
                    '只读': '否'
                },
                {
                    '地址': 1,
                    '点位名称': '当前温度',
                    '寄存器类型': '输入寄存器',
                    '初始值': 235,
                    '描述': '当前实际温度',
                    '单位': '0.1°C',
                    '最小值': 0,
                    '最大值': 1000,
                    '只读': '是'
                },
                {
                    '地址': 2,
                    '点位名称': '压力值',
                    '寄存器类型': '输入寄存器',
                    '初始值': 1013,
                    '描述': '当前压力值',
                    '单位': '0.1kPa',
                    '最小值': 0,
                    '最大值': 10000,
                    '只读': '是'
                }
            ]
            
            for row_idx, example in enumerate(examples, start=2):
                for col_idx, col_name in enumerate(self.COLUMNS, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=example.get(col_name, ''))
            
            # 自动调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # 添加说明工作表
            ws_info = wb.create_sheet("填写说明")
            instructions = [
                "寄存器点表配置说明",
                "",
                "1. 地址: Modbus 寄存器地址（0-65535）",
                "2. 点位名称: 点位的描述性名称",
                "3. 寄存器类型: 可选值为 '线圈'、'离散输入'、'保持寄存器'、'输入寄存器'",
                "   - 线圈(Coil): 可读写的位状态，对应功能码 01/05/15",
                "   - 离散输入(Discrete Input): 只读的位状态，对应功能码 02",
                "   - 保持寄存器(Holding Register): 可读写的16位寄存器，对应功能码 03/06/16",
                "   - 输入寄存器(Input Register): 只读的16位寄存器，对应功能码 04",
                "4. 初始值: 寄存器的初始值",
                "   - 线圈和离散输入: 0 或 1",
                "   - 寄存器: 0-65535",
                "5. 描述: 点位的详细描述",
                "6. 单位: 数值的单位（如 °C, kPa 等）",
                "7. 最小值: 允许的最小值（可选）",
                "8. 最大值: 允许的最大值（可选）",
                "9. 只读: 是否只读（'是' 或 '否'）",
                "",
                "注意事项:",
                "- 同一寄存器类型中，地址不能重复",
                "- 线圈和离散输入使用位地址",
                "- 保持寄存器和输入寄存器使用字地址",
                "- 只读点位只能通过界面或程序修改，不能通过 Modbus 写入"
            ]
            
            for row_idx, instruction in enumerate(instructions, start=1):
                cell = ws_info.cell(row=row_idx, column=1, value=instruction)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
            
            ws_info.column_dimensions['A'].width = 80
            
            # 保存文件
            wb.save(file_path)
            return OperationResult(True, data="模板创建成功")
            
        except Exception as e:
            return OperationResult(False, error=f"创建模板失败: {str(e)}")
    
    def validate_points(self, points: List[RegisterPoint]) -> OperationResult:
        """
        验证寄存器点位配置
        
        Args:
            points: 寄存器点位列表
            
        Returns:
            操作结果
        """
        errors = []
        
        # 检查地址重复
        address_map: Dict[str, List[int]] = {
            'coil': [],
            'discrete_input': [],
            'holding_register': [],
            'input_register': []
        }
        
        for point in points:
            if point.address in address_map[point.register_type]:
                errors.append(
                    f"地址冲突: {self.REGISTER_TYPES_CN[point.register_type]} "
                    f"地址 {point.address} 重复定义"
                )
            else:
                address_map[point.register_type].append(point.address)
            
            # 检查值的有效性
            if not point.validate_value(point.value):
                errors.append(
                    f"点位 '{point.name}' (地址 {point.address}): "
                    f"初始值 {point.value} 超出有效范围"
                )
        
        if errors:
            return OperationResult(False, error="\n".join(errors))
        
        return OperationResult(True, data=f"验证通过，共 {len(points)} 个点位")
    
    def group_points_by_type(self, points: List[RegisterPoint]) -> Dict[str, List[RegisterPoint]]:
        """
        按寄存器类型分组点位
        
        Args:
            points: 寄存器点位列表
            
        Returns:
            分组后的字典
        """
        grouped = {
            'coil': [],
            'discrete_input': [],
            'holding_register': [],
            'input_register': []
        }
        
        for point in points:
            grouped[point.register_type].append(point)
        
        # 按地址排序
        for key in grouped:
            grouped[key].sort(key=lambda p: p.address)
        
        return grouped
